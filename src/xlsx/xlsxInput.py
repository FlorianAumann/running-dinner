import openpyxl

from src.data.dinnerTeam import DinnerTeam
from src.data.participant import Participant


def read_teams_from_xlsx(file: str) -> [DinnerTeam]:
    """ Read in dinner teams from xlsx file
    The teams are expected in consecutive row starting at row 2
    The columns are expected in this order:\n
    Column 1: Address of the team\n
    Column 2: First name of first team member\n
    Column 3: Last name of first team member\n
    Column 4: Email of first team member\n
    Column 5: Phone number of first team member\n
    Column 6: Food restrictions of first team member (e.g. vegetarian or allergies)\n
    Column 2-6 may be repeated for an additional number of participants\n


    @param: file The file path to the xlsx file
    """
    try:
        wb = openpyxl.load_workbook(file)
    except Exception:
        return None

    if wb is None:
        return None
    ws = wb.worksheets[0]
    dinner_teams = []
    row_num = 2
    while ws.cell(row=row_num, column=2).value is not None:
        dt = DinnerTeam()
        dt.address = ws.cell(row=row_num, column=1).value
        dt.participants = []
        # Now get all participants for this team in the following columns
        column_num = 2
        while ws.cell(row=row_num, column=column_num).value is not None:
            participant = Participant()
            participant.firstName = str(ws.cell(row=row_num, column=column_num).value)
            participant.lastName = str(ws.cell(row=row_num, column=column_num + 1).value)
            participant.email = str(ws.cell(row=row_num, column=column_num + 2).value)
            participant.phone = str(ws.cell(row=row_num, column=column_num + 3).value)
            participant.food_restrictions = str(ws.cell(row=row_num, column=column_num + 4).value)
            dt.participants.append(participant)
            column_num += 5
        dinner_teams.append(dt)
        row_num += 1
    return dinner_teams
